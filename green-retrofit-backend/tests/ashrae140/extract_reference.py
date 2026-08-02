"""ASHRAE 140 5.2절 기준값을 BESTEST-GSR 배포본에서 추출해 CSV 로 고정한다.

xlsx 를 테스트 실행 시점에 읽지 않는 이유: 배포본이 갱신되면 기준값이 조용히
바뀌어 어제 통과하던 시험이 오늘 실패한다. **기준값은 커밋된 CSV 가 정본**이고,
이 스크립트는 그 CSV 를 어떻게 만들었는지 재현하기 위해서만 존재한다.

사용법:
    python extract_reference.py <BESTEST-GSR 체크아웃 경로>

출처:
  - results/resources/RESULTS5-2A.xlsx  표 B8-1(연간 난방) / B8-2(연간 현열 냉방)
  - results/historical/OpenStudio_3_11_0.csv  NREL 이 EnergyPlus 25.2.0 으로 낸 값
"""
import csv
import re
import sys
from pathlib import Path

# 표 B8-1 / B8-2 의 레이아웃 (RESULTS5-2A.xlsx, 시트 "Tables 1")
#   B열=케이스, C~H열=기준 프로그램 6종, I=Min, J=Max, K=Mean
TABLES = {
    "heating": {"title_row": 7, "first_data_row": 11, "label": "Table B8-1 Annual Heating Loads (MWh)"},
    "cooling": {"title_row": 59, "first_data_row": 63, "label": "Table B8-2 Annual Sensible Cooling Loads (MWh)"},
}
PROGRAM_COLS = {"BSIMAC": "C", "CSE": "D", "DeST": "E", "EnergyPlus": "F", "ESP-r": "G", "TRNSYS": "H"}


def extract(gsr_root: Path, out_dir: Path) -> None:
    import openpyxl

    xlsx = gsr_root / "results" / "resources" / "RESULTS5-2A.xlsx"
    ws = openpyxl.load_workbook(xlsx, data_only=True)["Tables 1"]

    rows = []
    for metric, spec in TABLES.items():
        row = spec["first_data_row"]
        while row <= ws.max_row:
            label = ws[f"B{row}"].value
            if not isinstance(label, str) or not re.match(r"^\d{3}\b", label.strip()):
                # 표 하나가 끝나면 케이스 번호로 시작하지 않는 행이 나온다
                if rows and rows[-1]["metric"] == metric:
                    break
                row += 1
                if row > spec["first_data_row"] + 80:
                    break
                continue
            case = label.strip().split()[0]
            lo, hi = ws[f"I{row}"].value, ws[f"J{row}"].value
            if isinstance(lo, (int, float)) and isinstance(hi, (int, float)):
                rows.append({
                    "case": case,
                    "metric": metric,
                    "unit": "MWh",
                    "ref_min": round(float(lo), 4),
                    "ref_max": round(float(hi), 4),
                    "ref_mean": round(float(ws[f"K{row}"].value), 4),
                    "energyplus_ref": round(float(ws[f"F{row}"].value), 4)
                    if isinstance(ws[f"F{row}"].value, (int, float)) else "",
                    "description": label.strip(),
                    "source": spec["label"],
                })
            row += 1

    out = out_dir / "std140_annual_loads.csv"
    with out.open("w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=list(rows[0].keys()))
        w.writeheader()
        w.writerows(rows)
    print(f"{out} — {len(rows)} 행")

    # NREL 이 우리와 같은 EnergyPlus 25.2.0 으로 낸 값. 범위 통과와 별개로
    # "같은 엔진끼리 얼마나 벌어지는가"를 보려면 이쪽이 훨씬 예민한 지표다.
    hist = gsr_root / "results" / "historical" / "OpenStudio_3_11_0.csv"
    ref = {}
    version = ""
    for line in hist.read_text(encoding="utf-8").splitlines():
        parts = line.split(",")
        if parts[0] == "program_name_and_version":
            version = parts[1]
        m = re.match(r"^Annual (Heating|Cooling) Loads (\d{3})$", parts[0])
        if m and len(parts) > 1:
            ref[(m.group(2), m.group(1).lower())] = float(parts[1])

    out2 = out_dir / "nrel_energyplus_25_2.csv"
    with out2.open("w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(["case", "metric", "unit", "value", "source"])
        for (case, metric), val in sorted(ref.items()):
            w.writerow([case, metric, "MWh", round(val, 4), f"BESTEST-GSR OpenStudio_3_11_0.csv ({version})"])
    print(f"{out2} — {len(ref)} 행 (기준 엔진: {version})")


def extract_programs(gsr_root, out_dir):
    """프로그램별 원값을 그대로 뽑는다 — **델타 범위를 직접 계산하기 위해서다.**

    델타 범위는 "각 프로그램의 (A−B)를 구한 뒤 그것들의 min/max" 여야 한다.
    min(A)−max(B) 처럼 집계값끼리 빼면 실제보다 훨씬 넓은 가짜 범위가 나온다.
    """
    import openpyxl
    xlsx = gsr_root / "results" / "resources" / "RESULTS5-2A.xlsx"
    ws = openpyxl.load_workbook(xlsx, data_only=True)["Tables 1"]

    rows = []
    for metric, spec in TABLES.items():
        for row in range(spec["first_data_row"], spec["first_data_row"] + 80):
            label = ws[f"B{row}"].value
            if not isinstance(label, str) or not re.match(r"^\d{3}\b", label.strip()):
                # ⚠️ 여기서 멈추지 않으면 난방 표(11행~)가 냉방 표(63행~)까지 넘어가
                # 난방 케이스에 냉방 값이 섞인다. 그러면 델타가 통째로 틀린다.
                break
            case = label.strip().split()[0]
            for prog, col in PROGRAM_COLS.items():
                v = ws[f"{col}{row}"].value
                if isinstance(v, (int, float)):
                    rows.append({"case": case, "metric": metric, "program": prog,
                                 "unit": "MWh", "value": round(float(v), 4)})

    out = out_dir / "std140_by_program.csv"
    with out.open("w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=["case", "metric", "program", "unit", "value"])
        w.writeheader()
        w.writerows(rows)
    print(f"{out} — {len(rows)} 행")


if __name__ == "__main__":
    if len(sys.argv) != 2:
        sys.exit(__doc__)
    _root, _out = Path(sys.argv[1]), Path(__file__).parent / "reference"
    extract(_root, _out)
    extract_programs(_root, _out)
